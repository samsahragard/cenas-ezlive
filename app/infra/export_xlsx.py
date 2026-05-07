# export google sheet
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
TOTAL_HEADER_FILL = PatternFill(fill_type="solid", fgColor="B8B0A8")
TOTAL_CELL_FILL = PatternFill(fill_type="solid", fgColor="EDE8E4")
BOLD_FONT = Font(bold=True)

def _safe_sheet_name(name: str) -> str:
    """
    Excel sheet names cannot contain: : \\ / ? * [ ]
    and must be <= 31 chars.
    """
    bad_chars = [":", "\\", "/", "?", "*", "[", "]"]
    cleaned = name
    for ch in bad_chars:
        cleaned = cleaned.replace(ch, "_")
    return cleaned[:31]

def _autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col_idx = column_cells[0].column

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))

        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

ORDER_TITLE_FONT = Font(bold=True, size=12)
ORDER_TITLE_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")


def _write_data_rows(ws, rows, col: dict[str, Any], start_row: int, num_value_cols: int) -> int:
    """
    Writes section-header + label/value rows for a single data column starting at start_row.
    Returns the next available row index.
    """
    current_row = start_row
    last_section = None

    for row_spec in rows:
        section = row_spec["section"]
        label = row_spec["label"]
        key = row_spec["key"]

        if section != last_section:
            ws.cell(row=current_row, column=1, value=section)
            ws.cell(row=current_row, column=1).font = BOLD_FONT
            ws.cell(row=current_row, column=1).fill = SECTION_FILL
            for col_idx in range(2, num_value_cols + 2):
                ws.cell(row=current_row, column=col_idx).fill = SECTION_FILL
            current_row += 1
            last_section = section

        label_cell = ws.cell(row=current_row, column=1, value=label)
        label_cell.alignment = Alignment(wrap_text=True, vertical="top")

        if col is not None:
            value = col["values"].get(key, "")
            value_cell = ws.cell(row=current_row, column=2, value=value)
            value_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

        current_row += 1

    return current_row


def _write_single_order_block(ws, rows, col: dict[str, Any], start_row: int) -> int:
    """
    Writes an individual order block (title + header + data rows).
    Returns the next available row index.
    """
    order_id = col["order_id"]

    # Filter to only rows that have a value for this order
    populated_rows = [r for r in rows if col["values"].get(r["key"], "") not in ("", None)]
    if not populated_rows:
        return start_row

    # Title
    title_cell = ws.cell(row=start_row, column=1, value=f"Order #{order_id}")
    title_cell.font = ORDER_TITLE_FONT
    title_cell.fill = ORDER_TITLE_FILL
    ws.cell(row=start_row, column=2).fill = ORDER_TITLE_FILL
    start_row += 1

    # Header
    for col_idx, header in enumerate(["Label", order_id], start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    start_row += 1

    return _write_data_rows(ws, populated_rows, col, start_row, num_value_cols=1)


def _write_view_sheet(ws, grid: dict[str, Any]) -> None:
    """
    Writes one grid into one worksheet.

    Layout:
    row 1: title
    row 2: blank
    row 3: column headers (Label + order ids)
    row 4+: section headers + row labels/values
    [blank rows]
    per-order individual blocks, one after another
    """
    view_name = str(grid["view"])
    rows = grid["rows"]
    columns = grid["columns"]

    # Title
    ws["A1"] = view_name.replace("_", " ").title()
    ws["A1"].font = Font(bold=True, size=14)

    # Header row
    header_row = 3
    ws.cell(row=header_row, column=1, value="Label")
    ws.cell(row=header_row, column=1).font = BOLD_FONT
    ws.cell(row=header_row, column=1).fill = HEADER_FILL
    ws.cell(row=header_row, column=1).alignment = Alignment(horizontal="center")

    for col_idx, col in enumerate(columns, start=2):
        is_total = col["order_id"] == "Total"
        cell = ws.cell(row=header_row, column=col_idx, value=col["order_id"])
        cell.font = BOLD_FONT
        cell.fill = TOTAL_HEADER_FILL if is_total else HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Main combined table
    current_row = header_row + 1
    last_section = None

    for row_spec in rows:
        section = row_spec["section"]
        label = row_spec["label"]
        key = row_spec["key"]

        if section != last_section:
            ws.cell(row=current_row, column=1, value=section)
            ws.cell(row=current_row, column=1).font = BOLD_FONT
            ws.cell(row=current_row, column=1).fill = SECTION_FILL
            for col_idx in range(2, len(columns) + 2):
                ws.cell(row=current_row, column=col_idx).fill = SECTION_FILL
            current_row += 1
            last_section = section

        label_cell = ws.cell(row=current_row, column=1, value=label)
        label_cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col_idx, col in enumerate(columns, start=2):
            is_total = col["order_id"] == "Total"
            value = col["values"].get(key, "")
            value_cell = ws.cell(row=current_row, column=col_idx, value=value)
            value_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
            if is_total and value:
                value_cell.font = BOLD_FONT
                value_cell.fill = TOTAL_CELL_FILL

        current_row += 1

    # Freeze top header + first column
    ws.freeze_panes = "B4"

    # Filter on main table only
    ws.auto_filter.ref = f"A3:{get_column_letter(len(columns) + 1)}{current_row - 1}"

    # Print settings: landscape, fit to 1 page wide, repeat header rows 1-3
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:3"

    # Individual order blocks (skip the synthetic Total column)
    current_row += 2  # blank separator
    for col in columns:
        if col["order_id"] == "Total":
            continue
        ws.row_breaks.append(Break(id=current_row - 1))
        current_row = _write_single_order_block(ws, rows, col, current_row)
        current_row += 2  # blank between orders

    _autosize_columns(ws)


def export_view_grids_to_xlsx(grids: dict[str, dict[str, Any]]) -> bytes:
    """
    Exports all view grids into a single workbook and returns the raw bytes.
    No disk I/O — safe for ephemeral deployments (e.g. Render).
    """
    wb = Workbook()

    # Remove default sheet; we'll add our own
    default_ws = wb.active
    wb.remove(default_ws)

    preferred_order = ["master", "kitchen", "driver", "prep_expo"]

    for view_name in preferred_order:
        if view_name not in grids:
            continue

        ws = wb.create_sheet(title=_safe_sheet_name(view_name.replace("_", " ").title()))
        _write_view_sheet(ws, grids[view_name])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
