"""Parse + ingest two ezCater Caterer Portal XLSX reports:

1. Order Data report  (parse_export_xlsx / apply_import) — financials per
   order: Food Total, Delivery Fee, Tip, Caterer Total Due, Driver name.

2. Delivery Performance Report  (parse_performance_report_xlsx /
   apply_performance_import, added 2026-05-10) — per-order tracking
   status: Tracked / Partially tracked / Untracked, plus the driver name
   as ezCater logged it, delivery start/complete times, and the
   On-time/Late result. Backs the per-driver payroll page; the
   tracking_status field gates the $10 / $1.50/mile / $5 bonuses Sam
   defined in 2026-05-10 spec.

Both files come from the ezCater portal export UI; both are uploaded by
Sam (or the manager) bi-weekly after each pay period.

Order Data sheet layout (as of 2026-05):
    Sheet 'Order Data', columns include:
        Order Number, Location, Street Address, City, State, Zip Code,
        Store Number, Store Name, Caterer Name, Event Date, Submitted At,
        Food Total, Promotion, Delivery Fee, Commission, Sales Tax,
        Sales Tax Remitted by ezCater, Tip, Payment Transaction Fee,
        Adjustments, Discounts, Misc Fees, Preferred Partner Program,
        Rewards, Caterer Total Due, Status, Order Paid By Caterer Payment
        Method, Source, Promotion Code, Driver, House Account ID, ...

Delivery Performance Report layout:
    Row 1 = title ("Delivery Performance Report from <start> to <end>")
    Row 2 = headers: Store # | Order # | Tracking | Driver | Event Date
                     | Delivery Start | Delivery Complete | Customer Event
                     Time | Result
    Rows 3..n = one delivery per row.
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime
from typing import IO

import openpyxl

from app.db import SessionLocal
from app.models import Order

logger = logging.getLogger(__name__)


# Store Number from xlsx -> internal store_id used by the rest of the codebase
_STORE_NUM_TO_ID = {1: "store_1", 2: "store_2", 3: "store_3", 4: "store_4"}
# store_id -> location key
_STORE_TO_LOCATION = {
    "store_1": "copperfield", "store_3": "copperfield",
    "store_2": "tomball",     "store_4": "tomball",
}


def _coerce_date(v) -> str | None:
    """The xlsx ships Event Date as a datetime when openpyxl decodes it.
    Normalize to ISO yyyy-mm-dd string (matches Order.delivery_date format)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    # Try a few common formats just in case
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_export_xlsx(stream: IO[bytes]) -> list[dict]:
    """Return one dict per row in the 'Order Data' sheet."""
    wb = openpyxl.load_workbook(stream, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.lower().strip() == "order data"),
                      wb.sheetnames[0])
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {name: header.index(name) for name in header if name}
    out = []
    required = ("Order Number", "Food Total", "Store Number")
    for col in required:
        if col not in idx:
            raise ValueError(f"Required column missing from xlsx: {col!r}. Found: {list(idx)[:8]}")
    for raw in rows[1:]:
        if raw is None:
            continue
        order_number = raw[idx["Order Number"]]
        if not order_number:
            continue
        # Skip the summary/total row at the bottom (Order Number = "Total")
        if str(order_number).strip().lower() == "total":
            continue
        food_total = raw[idx["Food Total"]]
        if food_total is None:
            continue
        store_num_v = raw[idx["Store Number"]]
        try:
            store_num = int(store_num_v) if store_num_v is not None else None
        except (TypeError, ValueError):
            store_num = None
        store_id = _STORE_NUM_TO_ID.get(store_num)
        location = _STORE_TO_LOCATION.get(store_id) if store_id else None
        def _money(col_name):
            if col_name not in idx:
                return None
            v = raw[idx[col_name]]
            try:
                return float(v) if v is not None else None
            except (TypeError, ValueError):
                return None

        def _txt(col_name):
            return (str(raw[idx[col_name]] or "").strip()
                    if col_name in idx and raw[idx[col_name]] is not None else "")
        street = _txt("Street Address")
        city = _txt("City")
        state = _txt("State")
        zipc = _txt("Zip Code")
        delivery_address = ", ".join(p for p in [street, city, f"{state} {zipc}".strip()] if p)
        rec = {
            "order_number": str(order_number).strip(),
            "food_total": float(food_total),
            "caterer_total_due": _money("Caterer Total Due"),
            "delivery_fee":      _money("Delivery Fee"),
            "tip_amount":        _money("Tip"),
            "ezcater_driver_name": (str(raw[idx["Driver"]] or "").strip()
                                    if "Driver" in idx and raw[idx["Driver"]] else None),
            "store_number": store_num,
            "store_id": store_id,
            "location": location,
            "pickup_kitchen": location,  # store_1/3 -> copperfield, store_2/4 -> tomball
            "delivery_address": delivery_address or None,
            "event_date": _coerce_date(raw[idx["Event Date"]]) if "Event Date" in idx else None,
            "submitted_at": (raw[idx["Submitted At"]].isoformat()
                             if "Submitted At" in idx and isinstance(raw[idx["Submitted At"]], datetime)
                             else None),
            "status": str(raw[idx["Status"]] or "").strip() if "Status" in idx else "",
            "source": str(raw[idx["Source"]] or "").strip() if "Source" in idx else "",
            "client": str(raw[idx["Location"]] or "").strip() if "Location" in idx else "",
        }
        out.append(rec)
    return out


def apply_import(rows: list[dict]) -> dict:
    """Match each parsed row to an existing Order by external_order_id and
    update Order.total_amount with the Food Total. For orders not yet in our
    DB, create a stub Order row so /reports/sales picks them up.

    Returns a summary: {parsed, updated, created, skipped, examples}.
    """
    db = SessionLocal()
    try:
        updated = 0
        created = 0
        skipped = 0
        examples_updated = []
        examples_created = []
        for r in rows:
            on = r["order_number"]
            existing = db.query(Order).filter(Order.external_order_id == on).first()
            if existing:
                old = existing.total_amount
                existing.total_amount = r["food_total"]
                # Populate the payroll/sales fields added in migration 11.
                # Always overwrite from the xlsx since it's the authoritative
                # ezCater-side data — re-uploading newer exports refreshes any
                # numbers that changed after a manual adjustment in the portal.
                if r.get("food_total") is not None:
                    existing.food_total = r["food_total"]
                if r.get("caterer_total_due") is not None:
                    existing.caterer_total_due = r["caterer_total_due"]
                if r.get("delivery_fee") is not None:
                    existing.delivery_fee = r["delivery_fee"]
                if r.get("tip_amount") is not None:
                    existing.tip_amount = r["tip_amount"]
                if r.get("ezcater_driver_name"):
                    existing.ezcater_driver_name = r["ezcater_driver_name"]
                if r.get("pickup_kitchen") and not existing.pickup_kitchen:
                    existing.pickup_kitchen = r["pickup_kitchen"]
                if r.get("delivery_address") and not existing.delivery_address:
                    existing.delivery_address = r["delivery_address"]
                if r.get("client") and not existing.client:
                    existing.client = r["client"]
                if r.get("event_date") and not existing.delivery_date:
                    existing.delivery_date = r["event_date"]
                if r.get("store_id") and not existing.origin_store_id:
                    existing.origin_store_id = r["store_id"]
                updated += 1
                if len(examples_updated) < 5:
                    examples_updated.append({"order_number": on, "old": old, "new": r["food_total"]})
            else:
                if not r.get("event_date") or not r.get("store_id"):
                    skipped += 1
                    continue
                stub = Order(
                    external_order_id=on,
                    client=r.get("client") or None,
                    delivery_address=r.get("delivery_address"),
                    delivery_date=r["event_date"],
                    origin_store_id=r["store_id"],
                    reported_store_id=r["store_id"],
                    total_amount=r["food_total"],
                    food_total=r.get("food_total"),
                    caterer_total_due=r.get("caterer_total_due"),
                    delivery_fee=r.get("delivery_fee"),
                    tip_amount=r.get("tip_amount"),
                    ezcater_driver_name=r.get("ezcater_driver_name"),
                    pickup_kitchen=r.get("pickup_kitchen"),
                    status=("cancelled" if "cancel" in r.get("status", "").lower()
                            else "imported"),
                )
                db.add(stub)
                created += 1
                if len(examples_created) < 5:
                    examples_created.append({"order_number": on, "amount": r["food_total"],
                                             "store": r["store_id"], "date": r["event_date"]})
        db.commit()
        return {
            "parsed":  len(rows),
            "updated": updated,
            "created": created,
            "skipped": skipped,
            "examples_updated": examples_updated,
            "examples_created": examples_created,
        }
    finally:
        db.close()


# ============================================================================
# Delivery Performance Report (separate xlsx export from ezCater portal).
# ============================================================================
# This is the file Sam sent on 2026-05-10 with one row per delivery and the
# Tracked / Partially tracked / Untracked status that gates the payroll
# bonuses. Joins to existing Orders on external_order_id (= "Order #" in
# the xlsx, after a dash is inserted into the 6-char code if missing).


def _format_time(v) -> str | None:
    """Performance Report ships times like 'Delivery Start' as openpyxl datetime
    objects. We render them as HH:MM AM/PM strings, the same format Sam's
    paycheck mock uses."""
    if v is None:
        return None
    if isinstance(v, (datetime,)):
        return v.strftime("%-I:%M %p").lstrip("0") if hasattr(v, "strftime") else v.isoformat()
    s = str(v).strip()
    return s or None


def _dash_order_number(raw: str) -> str:
    """Performance Report writes order numbers without the dash (e.g. '7TH22P'),
    but our Order.external_order_id is stored with a dash ('7TH-22P'). Insert
    the dash after position 3 for 6+ char codes that don't already have one."""
    raw = (raw or "").strip()
    if not raw or "-" in raw:
        return raw
    if len(raw) >= 4:
        return f"{raw[:3]}-{raw[3:]}"
    return raw


def parse_performance_report_xlsx(stream: IO[bytes]) -> list[dict]:
    """Parse the ezCater Delivery Performance Report xlsx export. Returns one
    dict per delivery row."""
    wb = openpyxl.load_workbook(stream, data_only=True)
    # Use the first sheet; ezCater's export only has one.
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []
    # Row 1 is the title ("Delivery Performance Report from..."); row 2 is the
    # actual column headers; data begins at row 3.
    header = [str(c).strip() if c is not None else "" for c in rows[1]]
    try:
        idx = {name: header.index(name) for name in
               ("Store #", "Order #", "Tracking", "Driver", "Event Date",
                "Delivery Start", "Delivery Complete", "Customer Event Time", "Result")}
    except ValueError as e:
        raise ValueError(f"Performance Report header doesn't match expected columns. "
                         f"Got: {header[:10]}") from e
    out = []
    for raw in rows[2:]:
        if raw is None or raw[idx["Order #"]] is None:
            continue
        order_no_raw = str(raw[idx["Order #"]]).strip()
        if not order_no_raw:
            continue
        store_num_v = raw[idx["Store #"]]
        try:
            store_num = int(store_num_v) if store_num_v is not None else None
        except (TypeError, ValueError):
            store_num = None
        store_id = _STORE_NUM_TO_ID.get(store_num)
        location = _STORE_TO_LOCATION.get(store_id) if store_id else None
        out.append({
            "order_number": _dash_order_number(order_no_raw),
            "store_number": store_num,
            "store_id": store_id,
            "pickup_kitchen": location,
            "tracking_status": str(raw[idx["Tracking"]] or "").strip() or None,
            "ezcater_driver_name": str(raw[idx["Driver"]] or "").strip() or None,
            "event_date": _coerce_date(raw[idx["Event Date"]]),
            "delivery_start_time": _format_time(raw[idx["Delivery Start"]]),
            "delivery_complete_time": _format_time(raw[idx["Delivery Complete"]]),
            "delivery_result": str(raw[idx["Result"]] or "").strip() or None,
        })
    return out


def apply_performance_import(rows: list[dict]) -> dict:
    """Match each parsed row to an existing Order by external_order_id and
    update the tracking + driver + delivery-time fields. Creates a stub Order
    if the order number isn't in our DB yet (same as apply_import).

    Returns: {parsed, updated, created, skipped, examples_updated}.
    """
    db = SessionLocal()
    try:
        updated = 0
        created = 0
        skipped = 0
        examples_updated = []
        for r in rows:
            on = r["order_number"]
            existing = db.query(Order).filter(Order.external_order_id == on).first()
            if existing:
                if r.get("tracking_status"):
                    existing.tracking_status = r["tracking_status"]
                if r.get("ezcater_driver_name"):
                    existing.ezcater_driver_name = r["ezcater_driver_name"]
                if r.get("pickup_kitchen") and not existing.pickup_kitchen:
                    existing.pickup_kitchen = r["pickup_kitchen"]
                if r.get("delivery_start_time"):
                    existing.delivery_start_time = r["delivery_start_time"]
                if r.get("delivery_complete_time"):
                    existing.delivery_complete_time = r["delivery_complete_time"]
                if r.get("delivery_result"):
                    existing.delivery_result = r["delivery_result"]
                if r.get("event_date") and not existing.delivery_date:
                    existing.delivery_date = r["event_date"]
                if r.get("store_id") and not existing.origin_store_id:
                    existing.origin_store_id = r["store_id"]
                updated += 1
                if len(examples_updated) < 5:
                    examples_updated.append({
                        "order_number": on,
                        "tracking": r.get("tracking_status"),
                        "driver": r.get("ezcater_driver_name"),
                    })
            else:
                if not r.get("event_date") or not r.get("store_id"):
                    skipped += 1
                    continue
                stub = Order(
                    external_order_id=on,
                    delivery_date=r["event_date"],
                    origin_store_id=r["store_id"],
                    reported_store_id=r["store_id"],
                    tracking_status=r.get("tracking_status"),
                    ezcater_driver_name=r.get("ezcater_driver_name"),
                    pickup_kitchen=r.get("pickup_kitchen"),
                    delivery_start_time=r.get("delivery_start_time"),
                    delivery_complete_time=r.get("delivery_complete_time"),
                    delivery_result=r.get("delivery_result"),
                    status="imported_performance",
                )
                db.add(stub)
                created += 1
        db.commit()
        return {
            "parsed":  len(rows),
            "updated": updated,
            "created": created,
            "skipped": skipped,
            # Different example-row shape than the Order Data importer (we have
            # tracking/driver, not old/new totals), and the result template
            # currently only knows how to render the Order Data shape — so
            # skip examples here and just surface the counts banner.
        }
    finally:
        db.close()
