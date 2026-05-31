"""Produce vendor email ingest — runs as a background thread inside the Flask app.

Polls orders@cenaskitchen.com every 60s. For each new email from an approved
sender (Alvarado / J. Luna / a Sam-forwarded JLUNA), downloads the attachment,
parses it (Claude vision for image, openpyxl for xlsx), maps vendor item names
to canonical names via aliases.json, and writes the result to
{PRODUCE_STATE_DIR}/{vendor}.json so the produce_order Blueprint shows the
fresh winning prices.

Gated by PRODUCE_INGEST_ENABLED=1 so dev environments don't pound IMAP.

Multi-worker safe: an fcntl file lock at {PRODUCE_STATE_DIR}/.ingest.lock
ensures only one process actually polls; other workers see the lock held and
no-op. (Render Starter is single-worker by default anyway.)

Env vars used:
  PRODUCE_INGEST_ENABLED   "1" to start the poller (default off)
  PRODUCE_STATE_DIR        where vendor JSONs + ingest state are written
                           (default <repo>/instance/produce)
  PRODUCE_CONFIG_DIR       where approved_senders.json + aliases.json live
                           (default <repo>/data/produce)
  ORDERS_EMAIL_PWD         SiteGround password for orders@cenaskitchen.com
                           (used for IMAP 993 — same as SMTP 465)
  ORDERS_IMAP_HOST         default gvam1078.siteground.biz
  ORDERS_IMAP_PORT         default 993
  ORDERS_IMAP_USER         default orders@cenaskitchen.com
  ANTHROPIC_API_KEY        for Claude vision (Alvarado image parsing)
  TELEGRAM_BOT_TOKEN       for Sam alerts on errors
  PRODUCE_TG_CHAT_ID       default 8612324971
"""
from __future__ import annotations

import base64
import email
import imaplib
import json
import logging
import os
import re
import threading
import time
import urllib.error
import urllib.request
from datetime import datetime, timezone
from email.header import decode_header
from pathlib import Path

logger = logging.getLogger(__name__)

# ============ Paths ============
REPO_ROOT = Path(__file__).resolve().parents[2]
CONFIG_DIR = Path(os.getenv("PRODUCE_CONFIG_DIR") or (REPO_ROOT / "data" / "produce"))
STATE_DIR = Path(os.getenv("PRODUCE_STATE_DIR") or (REPO_ROOT / "instance" / "produce"))

APPROVED_SENDERS_FILE = CONFIG_DIR / "approved_senders.json"
ALIASES_FILE = CONFIG_DIR / "aliases.json"

INGEST_STATE_FILE = STATE_DIR / "ingest_state.json"
LOCK_FILE = STATE_DIR / ".ingest.lock"
ATTACHMENT_DIR = STATE_DIR / "fetched"

# ============ IMAP config ============
IMAP_USER = os.getenv("ORDERS_IMAP_USER", "orders@cenaskitchen.com")
IMAP_HOST = os.getenv("ORDERS_IMAP_HOST", "gvam1078.siteground.biz")
IMAP_PORT = int(os.getenv("ORDERS_IMAP_PORT", "993"))
POLL_INTERVAL = int(os.getenv("PRODUCE_POLL_INTERVAL", "60"))

# Anthropic
ANTHROPIC_URL = "https://api.anthropic.com/v1/messages"
ANTHROPIC_MODEL = os.getenv("PRODUCE_VISION_MODEL", "claude-sonnet-4-6")
MAX_IMAGE_BYTES = 5 * 1024 * 1024

# Telegram
TG_API_BASE = "https://api.telegram.org"
TG_CHAT_ID = os.getenv("PRODUCE_TG_CHAT_ID", "8612324971")

_AICK_SECRETS = Path(r"C:\Users\sam\.openclaw\.secrets")


# ============ Secret resolvers ============
def _email_pwd() -> str:
    val = os.getenv("ORDERS_EMAIL_PWD")
    if val:
        return val.strip()
    f = _AICK_SECRETS / "orders_imap_pwd.txt"
    if f.exists():
        return f.read_text(encoding="utf-8").strip()
    raise RuntimeError("missing ORDERS_EMAIL_PWD env var and fallback file")


def _anthropic_key() -> str | None:
    val = os.getenv("ANTHROPIC_API_KEY")
    if val:
        return val.strip()
    f = _AICK_SECRETS / "anthropic_api_key.txt"
    if f.exists():
        return f.read_text(encoding="utf-8").strip()
    return None


def _tg_token() -> str | None:
    val = os.getenv("TELEGRAM_BOT_TOKEN")
    if val:
        return val.strip()
    f = _AICK_SECRETS / "ck_telegram_bot_token.txt"
    if f.exists():
        return f.read_text(encoding="utf-8").strip()
    return None


# ============ JSON helpers ============
def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _save_price_snapshots(vendor: str, payload: dict) -> None:
    """Insert one row per (vendor, item) into produce_price_snapshot.

    snapshot_date is parsed from `payload['date_range']` if present (the week
    the prices apply to — typically Mon-Sun); falls back to today's date.
    Uses INSERT OR IGNORE on the unique (snapshot_date, vendor, canonical_name,
    canonical_size) constraint so re-runs of the same email are idempotent."""
    from datetime import date
    import re
    from app.db import SessionLocal
    from app.models import ProducePriceSnapshot

    items = payload.get("items") or []
    if not items:
        return

    # Pick a snapshot_date: parse "5/5 - 5/11" / "5/5/2026 - 5/11/2026" style,
    # take the start date. Fall back to today.
    today_iso = date.today().isoformat()
    snapshot_date = today_iso
    dr = (payload.get("date_range") or "").strip()
    if dr:
        m = re.search(r"(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?", dr)
        if m:
            mo = int(m.group(1)); dy = int(m.group(2))
            yr_raw = m.group(3)
            yr = int(yr_raw) if yr_raw else date.today().year
            if yr < 100:
                yr += 2000
            try:
                snapshot_date = date(yr, mo, dy).isoformat()
            except ValueError:
                pass

    parsed_at = payload.get("parsed_at")
    from sqlalchemy.exc import IntegrityError as _IE
    db = SessionLocal()
    inserted = skipped = 0
    try:
        for it in items:
            cn = (it.get("canonical_name") or "").strip()
            cs = (it.get("canonical_size") or "").strip() or None
            price = it.get("price")
            if not cn or price is None:
                continue
            # Per-row commit so a race with another worker (or the bootstrap on
            # startup) doesn't roll back the whole batch — just skips the
            # conflicting row.
            db.add(ProducePriceSnapshot(
                snapshot_date=snapshot_date, vendor=vendor,
                canonical_name=cn, canonical_size=cs,
                price=float(price),
                raw_item_name=(it.get("vendor_name") or it.get("name")),
                parsed_at=parsed_at, date_range=dr or None,
            ))
            try:
                db.commit()
                inserted += 1
            except _IE:
                db.rollback()
                skipped += 1
    finally:
        db.close()
    logger.info("price-snapshot vendor=%s date=%s inserted=%d skipped=%d",
                vendor, snapshot_date, inserted, skipped)


def _read_json(path: Path, default=None):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return default if default is not None else {}
    return default if default is not None else {}


def _write_json(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, indent=2), encoding="utf-8")
    tmp.replace(path)


def _decode_h(s: str | None) -> str:
    if not s:
        return ""
    out = []
    for chunk, enc in decode_header(s):
        if isinstance(chunk, bytes):
            try:
                out.append(chunk.decode(enc or "utf-8", errors="replace"))
            except (LookupError, TypeError):
                out.append(chunk.decode("utf-8", errors="replace"))
        else:
            out.append(chunk)
    return "".join(out)


def _extract_email_address(from_header: str) -> str:
    """'Name <email@addr>' → 'email@addr', or returns the trimmed input."""
    m = re.search(r"<([^>]+)>", from_header or "")
    if m:
        return m.group(1).strip().lower()
    return (from_header or "").strip().lower()


# ============ Telegram ============
def _telegram(text: str) -> None:
    token = _tg_token()
    if not token:
        logger.info("(no telegram token) would have sent: %s", text[:200])
        return
    url = f"{TG_API_BASE}/bot{token}/sendMessage"
    payload = json.dumps({
        "chat_id": TG_CHAT_ID, "text": text, "disable_web_page_preview": True,
    }).encode("utf-8")
    req = urllib.request.Request(url, data=payload,
                                 headers={"Content-Type": "application/json"}, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            r.read()
    except Exception as e:
        logger.warning("telegram send failed: %s", e)


# ============ Image extractor (ported from produce_extract.py) ============
PROMPT = """You're parsing a produce vendor's price sheet (printed list, photographed). Extract EVERY row with a price.

Output ONLY valid JSON in this exact shape (no prose, no markdown fence):

{
  "date_range": "<the validity date range printed on the sheet, e.g. '05/03/2026-05/09/2026' or 'APRIL 27-30, 2026'. If not visible, use null>",
  "items": [
    {"vendor_name": "AGUACATE 32 CT # 1", "vendor_size": "32CT", "price": 36.00},
    {"vendor_name": "LIME 175 CT", "vendor_size": "175CT", "price": 48.00}
  ]
}

Rules:
- The sheet has multiple side-by-side columns. Extract items from ALL columns. There may be 50-80 items total.
- Item names may be Spanish (AGUACATE, LECHUGA, CEBOLLA, LIMON, CALABAZA, REPOLLO, ZANAHORIA, PAPA, TOMATE, CHILE, ELOTE, EJOTE, etc.) or English (LIME, ONION, EGGS, etc.) or mixed.
- "ESP" or "**ESP**" or "Especial" markers indicate a special - KEEP them in the vendor_name as written.
- For "N/A" prices: SKIP the item entirely. Do not include it in output.
- SKIP: header (vendor name/address/phone/dates), section labels (VERDURAS, FRUTAS, GROCERY, REFRESCOS, ESPECIES, CHILES SECOS, OTROS), footer/disclaimer text.
- vendor_size: just the size/pack/unit (e.g., "32CT", "50LBS", "CS", "BX", "EA", "head", "12CT", "33LB", "12oz"). Separate from the name even if they're written together.
- price: number only (36.00, NOT "$36.00").
- If a price is unclear or missing, omit that item.
- Be thorough. Don't summarize, don't skip categories.
"""


def _detect_media_type(filename: str) -> str:
    f = (filename or "").lower()
    if f.endswith((".jpg", ".jpeg")):
        return "image/jpeg"
    if f.endswith(".png"):
        return "image/png"
    if f.endswith(".gif"):
        return "image/gif"
    if f.endswith(".webp"):
        return "image/webp"
    return "image/jpeg"


def extract_from_image(image_bytes: bytes, filename: str) -> dict:
    api_key = _anthropic_key()
    if not api_key:
        raise RuntimeError("missing ANTHROPIC_API_KEY")
    if len(image_bytes) > MAX_IMAGE_BYTES:
        raise RuntimeError(f"image too large: {len(image_bytes)} bytes (max {MAX_IMAGE_BYTES})")

    media_type = _detect_media_type(filename)
    b64 = base64.standard_b64encode(image_bytes).decode("ascii")
    body = {
        "model": ANTHROPIC_MODEL,
        "max_tokens": 8000,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                {"type": "text", "text": PROMPT},
            ],
        }],
    }
    req = urllib.request.Request(
        ANTHROPIC_URL, data=json.dumps(body).encode("utf-8"),
        headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                 "content-type": "application/json"}, method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as r:
            resp = json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        body_text = e.read().decode("utf-8", errors="replace")[:500]
        raise RuntimeError(f"Anthropic HTTP {e.code}: {body_text}") from e

    text_parts = [c.get("text", "") for c in resp.get("content", []) if c.get("type") == "text"]
    if not text_parts:
        raise RuntimeError(f"no text in Anthropic response: {str(resp)[:300]}")
    full_text = "".join(text_parts).strip()
    if full_text.startswith("```"):
        lines = full_text.split("\n")
        if lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].startswith("```"):
            lines = lines[:-1]
        full_text = "\n".join(lines)
    try:
        return json.loads(full_text)
    except json.JSONDecodeError as e:
        raise RuntimeError(f"non-JSON from model: {e} -- snippet: {full_text[:300]}") from e


# ============ XLSX parser (ported from produce_parse_xlsx.py) ============
SKIP_NAME_PHRASES = (
    "PRICE", "ITEM", "FROM", "PHONE", "FAX", "PRODUCE COMPANY",
    "PRICES ARE", "INFORMATIONAL", "SUBJECT TO CHANGE",
)


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if 0 < float(v) < 999 else None
    s = str(v).strip().replace("$", "").replace(",", "")
    if not s:
        return None
    try:
        f = float(s)
        return f if 0 < f < 999 else None
    except ValueError:
        return None


def _looks_like_name(v) -> bool:
    if not isinstance(v, str):
        return False
    s = v.strip()
    if len(s) < 3:
        return False
    up = s.upper()
    for skip in SKIP_NAME_PHRASES:
        if skip in up:
            return False
    return bool(re.search(r"[A-Za-z]", s))


def _looks_like_size(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    return bool(s) and len(s) <= 20


def _parse_xlsx_date_range(rows_text: list[str]) -> str | None:
    blob = " ".join(rows_text)
    m = re.search(r"FROM[:\s]+([A-Z]+\.?\s*\d+)\s*[-–]\s*(\d+\s*[,.]?\s*\d{4})",
                  blob, re.IGNORECASE)
    if m:
        return f"{m.group(1).strip()} - {m.group(2).strip().rstrip('.')}"
    m = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})\s*[-–]\s*(\d{1,2}/\d{1,2}/\d{2,4})", blob)
    if m:
        return f"{m.group(1)} - {m.group(2)}"
    return None


def extract_from_xlsx(xlsx_path: Path) -> dict:
    from openpyxl import load_workbook  # local import keeps Flask startup light
    wb = load_workbook(xlsx_path, data_only=True)
    items = []
    rows_text = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            cells = list(row)
            rows_text.append(" | ".join(str(c) if c is not None else "" for c in cells))
            for col_idx, val in enumerate(cells):
                price = _to_float(val)
                if price is None or col_idx < 2:
                    continue
                size_v = cells[col_idx - 1]
                if not _looks_like_size(size_v):
                    continue
                name = None
                for offset in (3, 2):
                    if col_idx - offset >= 0:
                        cand = cells[col_idx - offset]
                        if _looks_like_name(cand):
                            name = cand.strip()
                            break
                if not name:
                    continue
                items.append({"name": name, "size": str(size_v).strip(), "price": price})

    seen = set()
    unique = []
    for it in items:
        key = (it["name"].upper(), it["size"].upper(), it["price"])
        if key not in seen:
            seen.add(key)
            unique.append(it)
    return {"date_range": _parse_xlsx_date_range(rows_text), "items": unique}


# ============ Alias mapping ============
_ALIAS_NORM_UNIT = re.compile(r"(\d)\s+(CT|LB|LBS|OZ|ONZ|DZ|EA|CS|BX|SK|PZ)\b")
_ALIAS_NORM_HASH = re.compile(r"#\s+(\d)")
_ALIAS_NORM_WS = re.compile(r"\s+")


def _normalize_alias_key(s: str) -> str:
    """Collapse whitespace + join digit-unit pairs ('32 CT' -> '32CT') and
    hash-digit pairs ('# 1' -> '#1') so the alias match is robust to parser
    output drift between Claude Vision runs / vendor formatting changes."""
    s = (s or "").upper().strip()
    s = _ALIAS_NORM_WS.sub(" ", s)
    s = _ALIAS_NORM_UNIT.sub(r"\1\2", s)
    s = _ALIAS_NORM_HASH.sub(r"#\1", s)
    return s


def _apply_aliases(vendor_key: str, items: list[dict]) -> tuple[list[dict], list[dict]]:
    """Returns (mapped_items, unmapped_vendor_items).
    mapped_items have canonical_name + canonical_size populated.

    Matching is normalized: 'AGUACATE 32 CT # 1|32CT' now matches the alias
    key 'AGUACATE 32CT #1|32CT'. The original alias-file keys also pass
    through the normalizer so old and new entries both work without
    file changes.
    """
    raw_aliases = _read_json(ALIASES_FILE).get(vendor_key, {})
    aliases = {_normalize_alias_key(k): v for k, v in raw_aliases.items()}
    mapped = []
    unmapped = []
    for it in items:
        v_name = (it.get("vendor_name") or it.get("name") or "").strip()
        v_size = (it.get("vendor_size") or it.get("size") or "").strip()
        key = _normalize_alias_key(f"{v_name}|{v_size}")
        alias = aliases.get(key)
        if alias:
            mapped.append({
                "canonical_name": alias["canonical_name"],
                "canonical_size": alias.get("canonical_size", ""),
                "vendor_name": v_name,
                "vendor_size": v_size,
                "price": it.get("price"),
            })
        else:
            unmapped.append({"vendor_name": v_name, "vendor_size": v_size, "price": it.get("price")})
    # Coverage warning: log when >30% of vendor lines couldn't be mapped — flags
    # alias drift early without waiting for Sam to notice blanks on the page.
    total = len(items)
    if total >= 20:
        unmapped_pct = len(unmapped) * 100 / total
        if unmapped_pct > 30:
            logger.warning(
                "alias coverage low: vendor=%s mapped=%d/%d unmapped=%d (%.0f%%) — "
                "aliases.json likely needs new entries",
                vendor_key, len(mapped), total, len(unmapped), unmapped_pct,
            )
    return mapped, unmapped


# ============ IMAP fetch ============
def _fetch_email_with_attachments(M, mid: bytes) -> tuple[email.message.Message, list[tuple[str, bytes]]]:
    """Fetch full email + return (parsed_msg, [(filename, bytes), ...])."""
    typ, fetched = M.fetch(mid, "(BODY.PEEK[])")
    if typ != "OK":
        raise RuntimeError(f"IMAP fetch failed: {typ}")
    msg = email.message_from_bytes(fetched[0][1])
    attachments = []
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        filename = part.get_filename()
        if not filename:
            continue
        filename = _decode_h(filename)
        payload = part.get_payload(decode=True)
        if payload:
            attachments.append((filename, payload))
    return msg, attachments


# ============ Vendor processing ============
def _process_email(M, mid: bytes, sender_info: dict) -> dict | None:
    """Returns a result dict {vendor, items_count, ...} or None on skip."""
    vendor = sender_info["vendor"]
    expected = sender_info["expected_format"]
    msg, attachments = _fetch_email_with_attachments(M, mid)
    subject = _decode_h(msg.get("Subject", ""))

    if not attachments:
        logger.info("mid=%s vendor=%s no attachments — skipping", mid, vendor)
        return None

    mid_str = mid.decode() if isinstance(mid, bytes) else str(mid)
    save_dir = ATTACHMENT_DIR / f"msg{mid_str}"
    save_dir.mkdir(parents=True, exist_ok=True)

    if expected == "image":
        chosen = next(((fn, b) for fn, b in attachments
                       if fn.lower().endswith((".jpg", ".jpeg", ".png", ".gif", ".webp"))), None)
        if not chosen:
            _telegram(f"⚠️ produce-ingest: {vendor} mid {mid_str} has no image attachment. "
                      f"Subject: {subject[:120]}. Skipping.")
            return None
        fn, b = chosen
        (save_dir / fn).write_bytes(b)
        result = extract_from_image(b, fn)
        items = result.get("items", [])
        date_range = result.get("date_range")
    elif expected == "xlsx":
        chosen = next(((fn, b) for fn, b in attachments
                       if fn.lower().endswith(".xlsx")), None)
        if not chosen:
            _telegram(f"⚠️ produce-ingest: {vendor} mid {mid_str} has no xlsx attachment. "
                      f"Subject: {subject[:120]}. Skipping.")
            return None
        fn, b = chosen
        path = save_dir / fn
        path.write_bytes(b)
        result = extract_from_xlsx(path)
        items = result.get("items", [])
        date_range = result.get("date_range")
    else:
        _telegram(f"⚠️ produce-ingest: unknown expected_format={expected!r} for sender {vendor}")
        return None

    if not items:
        _telegram(f"⚠️ produce-ingest: {vendor} mid {mid_str} parsed 0 items. "
                  f"Subject: {subject[:120]}. Format may have changed.")
        return None

    mapped, unmapped = _apply_aliases(vendor, items)
    vendor_file = STATE_DIR / f"{vendor}.json"
    payload = {
        "vendor": vendor,
        "date_range": date_range,
        "parsed_at": _now_iso(),
        "source_email_mid": mid_str,
        "subject": subject,
        "items": mapped,
        "unmapped_vendor_items": unmapped,
    }
    _write_json(vendor_file, payload)
    logger.info("mid=%s vendor=%s parsed=%d mapped=%d unmapped=%d",
                mid_str, vendor, len(items), len(mapped), len(unmapped))

    # Persist a snapshot row per item to produce_price_snapshot — feeds the
    # price-history view. Idempotent: re-runs of the same email won't dup.
    try:
        _save_price_snapshots(vendor, payload)
    except Exception:
        logger.exception("price-snapshot persistence failed for vendor=%s mid=%s", vendor, mid_str)

    if unmapped:
        sample = ", ".join(f"{u['vendor_name']} {u['vendor_size']}" for u in unmapped[:3])
        _telegram(f"🔤 produce-ingest: {vendor} has {len(unmapped)} items missing from "
                  f"aliases.json (sample: {sample}). Excluded from site this run.")

    return {"vendor": vendor, "items_count": len(mapped), "unmapped_count": len(unmapped),
            "mid": mid_str}


# ============ Polling loop ============
def _load_state() -> dict:
    return _read_json(INGEST_STATE_FILE, {"last_seen_mid": 0, "processed": {}})


def _save_state(state: dict) -> None:
    _write_json(INGEST_STATE_FILE, state)


def _poll_once() -> int:
    """Connect to IMAP, find new mail, process approved senders.
    Returns number of emails newly processed."""
    pwd = _email_pwd()
    senders = _read_json(APPROVED_SENDERS_FILE).get("senders", {})
    state = _load_state()
    last_seen = int(state.get("last_seen_mid", 0))
    processed_count = 0

    M = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    try:
        M.login(IMAP_USER, pwd)
        M.select("INBOX", readonly=True)
        typ, data = M.search(None, "ALL")
        if typ != "OK":
            return 0
        ids = data[0].split()
        if not ids:
            logger.info("poll: inbox empty")
            return 0
        max_mid = int(ids[-1])
        if last_seen == 0:
            # Baseline on first run — don't re-process the entire backlog.
            state["last_seen_mid"] = max_mid
            _save_state(state)
            logger.info("baseline last_seen_mid=%d", max_mid)
            return 0

        new_ids = [i for i in ids if int(i) > last_seen]
        logger.info("poll: inbox total=%d last_seen=%d max_mid=%d new=%d",
                    len(ids), last_seen, max_mid, len(new_ids))
        for nid in new_ids:
            mid_str = nid.decode()
            if mid_str in state.get("processed", {}):
                continue
            try:
                typ, hdr = M.fetch(nid, "(BODY.PEEK[HEADER.FIELDS (FROM SUBJECT)])")
                if typ != "OK":
                    continue
                msg_hdr = email.message_from_string(hdr[0][1].decode("utf-8", errors="replace"))
                from_hdr = _decode_h(msg_hdr.get("From", ""))
                addr = _extract_email_address(from_hdr)
                sender_info = senders.get(addr)
                logger.info("mid=%s from=%r addr=%r approved=%s (senders_loaded=%d)",
                            mid_str, from_hdr, addr, bool(sender_info), len(senders))
                if not sender_info:
                    # Unknown sender — skip silently. (The full skill flow Telegram'd Sam if
                    # there were attachments; we omit that here for simplicity.)
                    continue
                logger.info("mid=%s approved sender=%s vendor=%s",
                            mid_str, addr, sender_info["vendor"])
                result = _process_email(M, nid, sender_info)
                state.setdefault("processed", {})[mid_str] = {
                    "at": _now_iso(),
                    "vendor": sender_info["vendor"],
                    "items": (result or {}).get("items_count", 0),
                }
                if result:
                    processed_count += 1
            except Exception as e:
                logger.exception("error processing mid=%s", mid_str)
                _telegram(f"🚨 produce-ingest: error processing mid {mid_str}: {e}")
                state.setdefault("processed", {})[mid_str] = {
                    "at": _now_iso(), "error": str(e)[:300],
                }

        state["last_seen_mid"] = max(last_seen, max_mid)
        _save_state(state)
    finally:
        try: M.close()
        except Exception: pass
        try: M.logout()
        except Exception: pass

    return processed_count


# ============ Lock + thread loop ============
def _try_acquire_lock() -> object | None:
    """Best-effort cross-process lock. Returns the lock handle to keep open
    (and never close) on success, or None if another process holds it."""
    try:
        import fcntl  # POSIX only — Render runs Linux
    except ImportError:
        # Windows dev environment — skip the lock; assume single process.
        return object()
    LOCK_FILE.parent.mkdir(parents=True, exist_ok=True)
    fh = open(LOCK_FILE, "w")
    try:
        fcntl.flock(fh.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        fh.write(str(os.getpid()))
        fh.flush()
        return fh  # keep open for the lifetime of the process
    except OSError:
        fh.close()
        return None


def _loop() -> None:
    logger.info("produce-ingest poller starting (interval=%ds, host=%s, user=%s)",
                POLL_INTERVAL, IMAP_HOST, IMAP_USER)
    while True:
        try:
            n = _poll_once()
            if n > 0:
                logger.info("produce-ingest: processed %d email(s)", n)
        except Exception:
            logger.exception("poll iteration failed")
        time.sleep(POLL_INTERVAL)


def start_in_background() -> bool:
    """Called from create_app(). Starts the daemon thread iff
    PRODUCE_INGEST_ENABLED=1 and the cross-process lock is acquired.
    Returns True if the thread was started."""
    if os.getenv("PRODUCE_INGEST_ENABLED", "0") != "1":
        logger.info("produce-ingest disabled (set PRODUCE_INGEST_ENABLED=1 to enable)")
        return False
    lock = _try_acquire_lock()
    if lock is None:
        logger.info("produce-ingest lock held by another worker — skipping")
        return False
    t = threading.Thread(target=_loop, name="produce-ingest", daemon=True)
    t.start()
    return True


# ============ On-demand ingest (hourly cron + box failover entrypoint) ============
# The 60s in-process poller (above) is the real-time IMAP watcher and the SOLE writer
# of ingest_state.json. run_ingest_now() is the INDEPENDENT safety net the hourly
# Render cron + the box-failover Task Scheduler call. To stay race-free it does NOT
# call _poll_once (so it never touches ingest_state.json) - it only runs the idempotent
# "ensure each vendor's latest sheet is ingested" catch-up, which alone recovers a
# baseline-skip, a dead poller thread, or a disk reset. Serialized by its own NB lock
# (cron vs box). Writes a heartbeat + alerts on staleness. Sends NO vendor email, and
# re-parses (paid Claude vision) ONLY when a genuinely newer email exists.
HEARTBEAT_FILE = STATE_DIR / "ingest_heartbeat.json"
NOW_LOCK_FILE = STATE_DIR / ".ingest_now.lock"
STALE_DAYS = float(os.getenv("PRODUCE_STALE_DAYS", "12"))   # vendors send ~weekly + slack
SCAN_DEPTH = int(os.getenv("PRODUCE_CATCHUP_SCAN", "300"))  # cap the catch-up inbox walk
_NON_VENDOR_JSON = {"ingest_state", "ingest_heartbeat", "pending_orders",
                    "completed_orders"}


def _acquire_now_lock():
    """NB lock so the hourly cron + the box failover never run the catch-up
    concurrently (which would double-pay the Claude-vision parse). A SEPARATE file
    from the poller's lifetime lock, so it never deadlocks against it. Returns an open
    handle to keep until done, or None if another run already holds it."""
    try:
        import fcntl
    except ImportError:
        return object()  # non-POSIX dev/box: assume single runner
    NOW_LOCK_FILE.parent.mkdir(parents=True, exist_ok=True)
    fh = open(NOW_LOCK_FILE, "w")
    try:
        fcntl.flock(fh.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        return fh
    except OSError:
        fh.close()
        return None


def _release_now_lock(fh) -> None:
    try:
        import fcntl
        fcntl.flock(fh.fileno(), fcntl.LOCK_UN)
    except Exception:
        pass
    try:
        fh.close()
    except Exception:
        pass


def _age_days(parsed_at):
    """Days since parsed_at (tz-aware ISO). None if unparseable. total_seconds (not
    .days) so the threshold isn't floored; tolerant of a 'Z' suffix + naive stamps."""
    if not parsed_at:
        return None
    try:
        dt = datetime.fromisoformat(str(parsed_at).replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - dt).total_seconds() / 86400.0
    except Exception:
        return None


def _vendor_freshness() -> dict:
    """Per-vendor {parsed_at, items, age_days} from the {vendor}.json files. age_days
    is from parsed_at - and since the catch-up re-parses ONLY on a newer mid, parsed_at
    is a valid 'when we last got a fresh sheet' signal (retries don't reset it)."""
    out = {}
    try:
        files = list(STATE_DIR.glob("*.json"))
    except Exception:
        return out
    for vf in files:
        if vf.stem in _NON_VENDOR_JSON:
            continue
        doc = _read_json(vf, {})
        if "items" not in doc and "parsed_at" not in doc:
            continue
        out[vf.stem] = {"parsed_at": doc.get("parsed_at"),
                        "items": len(doc.get("items", [])),
                        "age_days": _age_days(doc.get("parsed_at"))}
    return out


def _ensure_latest_per_vendor(attempted=None):
    """Catch-up: guarantee each vendor's NEWEST approved email is reflected in
    {vendor}.json. Fixes the baseline-skip / dead-poller / disk-reset gap. Idempotent:
    re-parses ONLY when a newer email exists than what's stored (source_email_mid), and
    NEVER re-parses the same mid twice (the `attempted` guard) - so a newest email that
    legitimately yields 0 items (and so writes no file) can't loop a paid vision parse
    every hour. Returns (result, attempted). `attempted` = {vendor: last_mid_tried},
    carried in the heartbeat."""
    attempted = dict(attempted or {})
    senders = _read_json(APPROVED_SENDERS_FILE).get("senders", {})
    if not senders:
        return {}, attempted
    pwd = _email_pwd()
    want = {s["vendor"] for s in senders.values() if s.get("vendor")}
    result: dict = {}
    M = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    try:
        M.login(IMAP_USER, pwd)
        M.select("INBOX", readonly=True)
        typ, data = M.search(None, "ALL")
        if typ != "OK":
            return result, attempted
        ids = data[0].split()
        newest: dict = {}  # vendor -> (mid, sender_info)
        for nid in reversed(ids[-SCAN_DEPTH:]):   # cap the walk; newest-first
            if len(newest) >= len(want):
                break
            typ, hdr = M.fetch(nid, "(BODY.PEEK[HEADER.FIELDS (FROM)])")
            if typ != "OK":
                continue
            msg_hdr = email.message_from_string(hdr[0][1].decode("utf-8", "replace"))
            addr = _extract_email_address(_decode_h(msg_hdr.get("From", "")))
            si = senders.get(addr)
            if si and si.get("vendor") and si["vendor"] not in newest:
                newest[si["vendor"]] = (nid, si)
        for vendor, (nid, si) in newest.items():
            mid_str = nid.decode() if isinstance(nid, bytes) else str(nid)
            cur = _read_json(STATE_DIR / f"{vendor}.json", {})
            if cur.get("source_email_mid") == mid_str and cur.get("items"):
                result[vendor] = "fresh"
                continue
            if attempted.get(vendor) == mid_str:
                result[vendor] = f"skip (already attempted mid {mid_str})"
                continue
            attempted[vendor] = mid_str   # record BEFORE the parse so a failure can't loop
            try:
                r = _process_email(M, nid, si)
                result[vendor] = f"ingested mid {mid_str} items={(r or {}).get('items_count', 0)}"
            except Exception as e:  # noqa: BLE001
                logger.exception("catch-up failed vendor=%s mid=%s", vendor, mid_str)
                result[vendor] = f"error: {str(e)[:120]}"
    finally:
        try:
            M.close()
        except Exception:
            pass
        try:
            M.logout()
        except Exception:
            pass
    return result, attempted


def _maybe_alert_stale(fresh: dict, prev_alert_at=None) -> str | None:
    """Telegram Sam at most once / 24h if a vendor has NO priced sheet at all, or its
    latest sheet is older than STALE_DAYS. Low-frequency, low-false-positive - the
    'it went quiet' signal so ingest can't silently die for weeks again. Called under
    run_ingest_now's lock with prev_alert_at passed in, so no concurrent-alert race."""
    if prev_alert_at:
        try:
            if (datetime.now(timezone.utc) - datetime.fromisoformat(str(prev_alert_at).replace("Z", "+00:00"))).total_seconds() < 86400:
                return None
        except Exception:
            pass
    senders = _read_json(APPROVED_SENDERS_FILE).get("senders", {})
    want = {s["vendor"] for s in senders.values() if s.get("vendor")}
    problems = []
    for v in sorted(want):
        f = fresh.get(v)
        if not f or not f.get("items"):
            problems.append(f"{v}: NO prices")
        elif f.get("age_days") is not None and f["age_days"] > STALE_DAYS:
            problems.append(f"{v}: {f['age_days']:.0f}d old")
    if not problems:
        return None
    _telegram("⚠ produce-ingest staleness: " + "; ".join(problems)
              + ". Check orders@ + the poller/cron.")
    return "; ".join(problems)


def run_ingest_now() -> dict:
    """On-demand catch-up for the hourly Render cron + the box failover. NB-locked
    (cron vs box). Does NOT call _poll_once - the 60s poller owns real-time polling +
    ingest_state.json; this only ensures each vendor's latest sheet is ingested (which
    alone recovers a dead poller / baseline-skip / disk reset), writes a heartbeat, and
    alerts on staleness. Idempotent; sends no vendor email."""
    lock = _acquire_now_lock()
    if lock is None:
        return {"skipped": "another ingest run in progress"}
    summary: dict = {"catch_up": {}, "alert": None}
    try:
        prev = _read_json(HEARTBEAT_FILE, {})
        attempted = prev.get("attempted", {})
        try:
            summary["catch_up"], attempted = _ensure_latest_per_vendor(attempted)
        except Exception as e:  # noqa: BLE001
            logger.exception("run_ingest_now catch-up failed")
            summary["catch_up_error"] = str(e)[:160]
        try:
            fresh = _vendor_freshness()
            summary["freshness"] = fresh
            summary["alert"] = _maybe_alert_stale(fresh, prev.get("last_alert_at"))
            hb = {"at": _now_iso(), "freshness": fresh, "attempted": attempted,
                  "last_alert_at": _now_iso() if summary.get("alert") else prev.get("last_alert_at")}
            _write_json(HEARTBEAT_FILE, hb)
        except Exception:
            logger.exception("run_ingest_now freshness/heartbeat failed")
        return summary
    finally:
        _release_now_lock(lock)
